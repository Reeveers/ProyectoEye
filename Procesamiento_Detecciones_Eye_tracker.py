import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import math
import matplotlib.pyplot as plt
from sklearn.cluster import DBSCAN

#-------------VARIABLES GLOBALES-------------
NumFijaciones = 0
umbral_regresion = 20  # Ajustar el umbral para las regresiones
epsilon = 1e-5  # Umbral para evitar errores de precisión pequeños
umbral_sacada = 100  # Distancia mínima para considerar una sacada
duracion_minima_fijacion = 100  # Mínimo tiempo en ms para considerar una fijación

def calcular_distancia(punto1, punto2):
    return math.sqrt((punto2[0] - punto1[0]) ** 2 + (punto2[1] - punto1[1]) ** 2)

def calculate_angular_velocity(x1, y1, x2, y2, delta_t):
    if delta_t == 0 or (abs(x2 - x1) < epsilon and abs(y2 - y1) < epsilon):
        return 0  # Evitar la división por cero y valores muy pequeños
    delta_x = x2 - x1
    delta_y = y2 - y1
    return math.atan2(delta_y, delta_x) / delta_t

def calcular_angulo(x1, y1, x2, y2, x3, y3):
    a = calcular_distancia((x2, y2), (x3, y3))
    b = calcular_distancia((x1, y1), (x3, y3))
    c = calcular_distancia((x1, y1), (x2, y2))
    if b == 0 or c == 0:
        return 0
    cos_angulo = (b**2 + c**2 - a**2) / (2 * b * c)
    cos_angulo = max(-1, min(1, cos_angulo))
    angulo = math.degrees(math.acos(cos_angulo))
    return angulo

# Filtro de datos para suavizar el ruido
def filtrar_datos(cursor_history, umbral_distancia=10):
    datos_filtrados = [cursor_history[0]]  # Iniciar con el primer punto
    for i in range(1, len(cursor_history)):
        if calcular_distancia(cursor_history[i]['position'], cursor_history[i-1]['position']) > umbral_distancia:
            datos_filtrados.append(cursor_history[i])
    return datos_filtrados

# Detectar fijaciones usando DBSCAN y duración mínima
def detectar_fijaciones_dbscan(cursor_history, eps=50, min_samples=5):
    puntos = [(p['position'][0], p['position'][1]) for p in cursor_history]
    tiempos = [p['time'] for p in cursor_history]
    clustering = DBSCAN(eps=eps, min_samples=min_samples).fit(puntos)
    etiquetas = clustering.labels_

    fijaciones = []
    for i, etiqueta in enumerate(etiquetas):
        if etiqueta != -1:  # -1 indica que es ruido
            # Verificar si el tiempo dentro del clúster es suficiente para considerarse fijación
            cluster_puntos = [cursor_history[j] for j, et in enumerate(etiquetas) if et == etiqueta]
            duracion = cluster_puntos[-1]['time'] - cluster_puntos[0]['time']
            if duracion >= duracion_minima_fijacion:
                fijaciones.extend(cluster_puntos)
    return fijaciones

# Detección de sacadas
def detectar_sacadas(cursor_history, umbral_sacada):
    sacadas = []
    for i in range(1, len(cursor_history)):
        distancia = calcular_distancia(cursor_history[i]['position'], cursor_history[i-1]['position'])
        if distancia > umbral_sacada:
            sacadas.append(cursor_history[i])
    return sacadas

def procesar_fijaciones(cursor_history):
    global NumFijaciones
    fijaciones = detectar_fijaciones_dbscan(cursor_history)
    NumFijaciones = len(fijaciones)
    return fijaciones, cursor_history

def deteccion_regresiones(cursor_history, umbral_aceleracion, umbral_angulo):
    regresiones = []
    for i in range(2, len(cursor_history)):
        before_previous_packet = cursor_history[i - 2]
        previous_packet = cursor_history[i - 1]
        current_packet = cursor_history[i]
        
        if current_packet['position'][0] < previous_packet['position'][0]:
            delta_t = current_packet['time'] - previous_packet['time']
            angular_velocity = calculate_angular_velocity(
                previous_packet['position'][0], previous_packet['position'][1],
                current_packet['position'][0], current_packet['position'][1],
                delta_t
            )
            angulo = calcular_angulo(
                before_previous_packet['position'][0], before_previous_packet['position'][1],
                previous_packet['position'][0], previous_packet['position'][1],
                current_packet['position'][0], current_packet['position'][1]
            )
            
            if abs(angular_velocity) < umbral_aceleracion and angulo >= umbral_angulo:
                regresiones.append(current_packet)
    
    return regresiones

# Resaltar fijaciones en Excel
def resaltar_fijaciones_excel(cursor_history, output_file, fijaciones_indices):
    df = pd.DataFrame(cursor_history)
    df.to_excel(output_file, index=False)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        fixation_cell = row[df.columns.get_loc("fixation")]
        if fixation_cell.value == 1.0:
            for cell in row:
                cell.fill = yellow_fill

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for idx in fijaciones_indices:
        for cell in sheet.iter_rows(min_row=idx+2, max_row=idx+2, min_col=1, max_col=sheet.max_column):
            for c in cell:
                c.fill = red_fill
    
    workbook.save(output_file)

# Leer el archivo JSON
with open('cursor_history.json', 'r') as json_file:
    cursor_history = json.load(json_file)

print(f"-------------------------------DETECCIONES------------------------------")

# Filtrar los datos para reducir el ruido
cursor_history = filtrar_datos(cursor_history)

# Procesar fijaciones
fijaciones, cursor_history = procesar_fijaciones(cursor_history)
print(f"Fijaciones: {NumFijaciones}")
resaltar_fijaciones_excel(cursor_history, "cursor_history_test.xlsx", [cursor_history.index(f) for f in fijaciones])

# Detectar sacadas
sacadas = detectar_sacadas(cursor_history, umbral_sacada)
print(f"Sacadas: {len(sacadas)}")

# Detectar regresiones
umbral_aceleracion = 0.05  # Ajuste de umbral para mejor detección de regresiones
umbral_angulo = 5
regresiones = deteccion_regresiones(cursor_history, umbral_aceleracion, umbral_angulo)
print(f"Regresiones: {len(regresiones)}")

# Extraer coordenadas
x_coords = [event['position'][0] for event in cursor_history]
y_coords = [event['position'][1] for event in cursor_history]
x_fijaciones = [f['position'][0] for f in fijaciones]
y_fijaciones = [f['position'][1] for f in fijaciones]
x_regresiones = [event['position'][0] for event in regresiones]
y_regresiones = [event['position'][1] for event in regresiones]
x_sacadas = [event['position'][0] for event in sacadas]
y_sacadas = [event['position'][1] for event in sacadas]

# Crear la gráfica
plt.figure(figsize=(10, 6))
plt.scatter(x_coords, y_coords, c='blue', marker='o', label='Puntos')
plt.scatter(x_fijaciones, y_fijaciones, c='red', marker='x', s=100, label='Fijaciones')
plt.scatter(x_regresiones, y_regresiones, c='green', marker='*', s=100, label='Regresiones')
plt.scatter(x_sacadas, y_sacadas, c='orange', marker='^', s=100, label='Sacadas')
plt.title('Gráfica de Detecciones')
plt.xlim(0, 1980)
plt.ylim(0, 1080)
plt.xlabel('Anchura Pantalla X')
plt.ylabel('Altura Pantalla Y')
plt.legend()
plt.grid(True)
plt.show()
